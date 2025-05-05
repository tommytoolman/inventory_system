"""
Provides an interactive CLI menu to use the ProductMatcher class (imported from app/tools/). 
Allows users to find matches between platforms, review them one by one, confirm/reject matches, save/load progress, 
commit confirmed merges to the database, and export matches to Excel. 
This seems to be the main user interface for the matching logic.
"""

import os, sys
import asyncio
import logging

from dotenv import load_dotenv

from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
from sqlalchemy.orm import sessionmaker

from app.tools.product_matcher import ProductMatcher

# Set up logging
logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)

load_dotenv()

def show_pending_matches(confirmed_matches):
    """Display matches waiting to be committed"""
    if not confirmed_matches:
        print("No confirmed matches pending database commit.")
        return
        
    print(f"\nYou have {len(confirmed_matches)} confirmed matches waiting to be committed:")
    
    for i, match in enumerate(confirmed_matches):
        # Get the platform names from the match object
        platforms = match.get('platforms', [])
        if not platforms:
            # If platforms not stored, extract from keys
            platforms = [k.replace('_product', '') for k in match.keys() if k.endswith('_product')]
        
        print(f"\nMatch {i+1}:")
        for platform in platforms:
            if f'{platform}_product' in match:
                product = match[f'{platform}_product']
                price = f"£{product['price']:,.0f}" if product.get('price') else "No price"
                print(f"  {platform.upper()}: {product['sku']} - {product['title']} - {price}")
    
    print("\nUse 'Commit confirmed matches' to apply these changes to the database.")

async def find_and_review_matches(matcher, confirmed_matches, processed_pairs):
    """Find and review new matches"""
    try:
        # Get available platforms
        platforms_data = await matcher._get_products_by_platform()
        available_platforms = list(platforms_data.keys())
        
        if len(available_platforms) < 2:
            print("Error: Need at least 2 platforms with products to perform matching")
            return
            
        print(f"Available platforms: {', '.join(available_platforms)}")
        
        # Choose platforms to compare
        platform1 = input(f"First platform (default: {available_platforms[0]}): ").lower() or available_platforms[0]
        platform2 = input(f"Second platform (default: {available_platforms[1]}): ").lower() or available_platforms[1]
        
        if platform1 not in available_platforms or platform2 not in available_platforms:
            print(f"Error: One or both platforms not available. Available platforms: {', '.join(available_platforms)}")
            return
            
        # Set minimum confidence threshold
        min_confidence = int(input("Minimum confidence threshold (1-100, default: 85): ") or "85")
        
        # Find potential matches
        print(f"Finding potential matches between {platform1} and {platform2}...")
        matches = await matcher.find_potential_matches(min_confidence=min_confidence, 
                                                 platform1=platform1, 
                                                 platform2=platform2)
        
        # Filter out already processed matches
        new_matches = []
        for match in matches:
            match_key = (match[f'{platform1}_product']['id'], match[f'{platform2}_product']['id'])
            if match_key not in processed_pairs:
                new_matches.append(match)
        
        if not new_matches:
            print("No new potential matches found!")
            
            # Ask if the user wants to process any pending confirmed matches
            if confirmed_matches:
                choice = input(f"You have {len(confirmed_matches)} confirmed matches waiting to be committed. Process them now? (y/n): ").lower()
                if choice == 'y':
                    print(f"\nCommitting {len(confirmed_matches)} confirmed matches...")
                    merged = await matcher.merge_products(confirmed_matches)
                    print(f"Successfully merged {merged} products.")
                    # Clear confirmed matches after processing
                    confirmed_matches.clear()
                    await matcher.save_progress(confirmed_matches, processed_pairs)
            
            # Ask if they want to reset processed pairs
            choice = input("Would you like to reset all processed pairs to review matches again? (y/n): ").lower()
            if choice == 'y':
                processed_pairs.clear()
                await matcher.save_progress(confirmed_matches, processed_pairs)
                print("All processed pairs have been reset. Run the matcher again to review matches.")
            
            return
        
        print(f"Found {len(new_matches)} potential matches to review.")
        
        # Interactive review
        i = 0
        while i < len(new_matches):
            match = new_matches[i]
            product1 = match[f'{platform1}_product']
            product2 = match[f'{platform2}_product']
            confidence = match['confidence']
            
            print("\n" + "="*80)
            print(f"Match {i+1}/{len(new_matches)} (Confidence: {confidence}%)")
            print("-"*80)
            
            # Display match details
            price1 = f"£{product1['price']:,.0f}" if product1.get('price') else "No price"
            print(f"{platform1.upper()}: [{product1['sku']}] {product1['title']} - {price1}")
            print(f"Brand: {product1['brand']}, Model: {product1['model']}")
            
            print("-"*80)
            
            price2 = f"£{product2['price']:,.0f}" if product2.get('price') else "No price"
            print(f"{platform2.upper()}: [{product2['sku']}] {product2['title']} - {price2}")
            print(f"Brand: {product2['brand']}, Model: {product2['model']}")
            
            print("="*80)
            
            # Expanded options for input
            print("Options: y (confirm), n (reject), s (save & quit), b (back), q (quit without saving)")
            choice = input("Your choice: ").lower()
            
            # Mark as processed regardless of choice
            prod1_id = product1['id']
            prod2_id = product2['id']
            
            if choice == 's':
                # Save progress and quit
                processed_pairs.add((prod1_id, prod2_id))
                await matcher.save_progress(confirmed_matches, processed_pairs)
                print("Progress saved. You can resume later.")
                return
            elif choice == 'q':
                # Quit without saving this pair
                print("Exiting without saving current match progress.")
                return
            elif choice == 'b' and i > 0:
                # Go back to previous match (if not the first one)
                i -= 1
                # Remove the last match from processed pairs if it was processed
                last_match = new_matches[i]
                last_key = (last_match[f'{platform1}_product']['id'], last_match[f'{platform2}_product']['id'])
                if last_key in processed_pairs:
                    processed_pairs.remove(last_key)
                continue
            elif choice == 'y':
                # Confirm match
                processed_pairs.add((prod1_id, prod2_id))
                match['platforms'] = [platform1, platform2]  # Store platforms for later use
                confirmed_matches.append(match)
                print(f"Match confirmed! ({len(confirmed_matches)} total confirmed)")
                # Save after each confirmation
                await matcher.save_progress(confirmed_matches, processed_pairs)
            else:
                # Default to rejecting
                processed_pairs.add((prod1_id, prod2_id))
                print("Match rejected.")
                await matcher.save_progress(confirmed_matches, processed_pairs)
            
            # Move to next match
            i += 1
        
        # After reviewing all matches
        print("\nFinished reviewing all potential matches.")
        if confirmed_matches:
            print(f"You have {len(confirmed_matches)} confirmed matches waiting to be committed.")
            if input("Would you like to commit these matches now? (y/n): ").lower() == 'y':
                print(f"\nCommitting {len(confirmed_matches)} confirmed matches...")
                merged = await matcher.merge_products(confirmed_matches)
                print(f"Successfully merged {merged} products.")
                # Clear confirmed matches after processing
                confirmed_matches.clear()
                await matcher.save_progress(confirmed_matches, processed_pairs)
                
    except Exception as e:
        logger.error(f"Error in find_and_review_matches: {str(e)}")
        import traceback
        traceback.print_exc()
        print(f"An error occurred: {str(e)}")

async def edit_confirmed_matches(matcher, confirmed_matches):
    """Review and edit previously confirmed matches"""
    if not confirmed_matches:
        print("No confirmed matches to edit.")
        return
    
    while True:
        # Display the confirmed matches
        show_pending_matches(confirmed_matches)
        
        print("\nEdit Options:")
        print("  number: Select a match to remove (e.g., '3' to remove match #3)")
        print("  c: Clear all confirmed matches")
        print("  q: Return to main menu")
        
        choice = input("\nYour choice: ").lower()
        
        if choice == 'q':
            break
        elif choice == 'c':
            if input("Are you sure you want to clear ALL confirmed matches? (y/n): ").lower() == 'y':
                confirmed_matches.clear()
                await matcher.save_progress(confirmed_matches, set())
                print("All confirmed matches cleared.")
                break
        else:
            try:
                idx = int(choice) - 1
                if 0 <= idx < len(confirmed_matches):
                    match = confirmed_matches[idx]
                    
                    # Print match details
                    platforms = match.get('platforms', [])
                    if not platforms:
                        platforms = [k.replace('_product', '') for k in match.keys() if k.endswith('_product')]
                    
                    print(f"\nRemoving match {idx+1}:")
                    for platform in platforms:
                        if f'{platform}_product' in match:
                            product = match[f'{platform}_product']
                            print(f"  {platform.upper()}: {product['sku']} - {product['title']}")
                    
                    if input("Are you sure you want to remove this match? (y/n): ").lower() == 'y':
                        confirmed_matches.pop(idx)
                        await matcher.save_progress(confirmed_matches, set())
                        print("Match removed.")
                else:
                    print("Invalid match number.")
            except ValueError:
                print("Invalid input. Please enter a number, 'c', or 'q'.")

async def export_matches_to_excel(matcher):
    """Export potential matches to Excel for review"""
    import pandas as pd
    from datetime import datetime, timezone
    from openpyxl.formatting.rule import ColorScaleRule
    
    # Get available platforms
    platforms_data = await matcher._get_products_by_platform()
    available_platforms = list(platforms_data.keys())
    
    print(f"Available platforms: {', '.join(available_platforms)}")
    
    # Choose platforms to compare
    platform1 = input(f"First platform (default: {available_platforms[0]}): ").lower() or available_platforms[0]
    platform2 = input(f"Second platform (default: {available_platforms[1]}): ").lower() or available_platforms[1]
    
    # Set minimum confidence threshold
    min_confidence = int(input("Minimum confidence threshold (1-100, default: 70): ") or "70")
    
    # Find all potential matches (use a lower threshold for export)
    print(f"Finding potential matches between {platform1} and {platform2}...")
    matches = await matcher.find_potential_matches(min_confidence=min_confidence, 
                                            platform1=platform1, 
                                            platform2=platform2)
    
    if not matches:
        print("No matches found with the current threshold.")
        return
    
    # Create data for export
    data = []
    for match in matches:
        product1 = match[f'{platform1}_product']
        product2 = match[f'{platform2}_product']
        confidence = match['confidence']
        
        # Calculate price difference
        price1 = float(product1.get('price', 0) or 0)
        price2 = float(product2.get('price', 0) or 0)
        
        price_diff = 0
        price_diff_pct = 0
        
        if price1 > 0 and price2 > 0:
            price_diff = price1 - price2
            if price1 >= price2:
                price_diff_pct = (price1 - price2) / price1 * 100
            else:
                price_diff_pct = (price2 - price1) / price2 * 100
        
        data.append({
            f"{platform1}_sku": product1['sku'],
            f"{platform1}_title": product1['title'],
            f"{platform1}_brand": product1['brand'],
            f"{platform1}_model": product1['model'],
            f"{platform1}_price": price1,
            f"{platform2}_sku": product2['sku'],
            f"{platform2}_title": product2['title'],
            f"{platform2}_brand": product2['brand'],
            f"{platform2}_model": product2['model'],
            f"{platform2}_price": price2,
            "price_difference": price_diff,
            "price_diff_pct": price_diff_pct,
            "confidence": confidence
        })
    
    # Create dataframe and export to Excel
    df = pd.DataFrame(data)
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"product_matches_{platform1}_vs_{platform2}_{timestamp}.xlsx"
    
    # Export to Excel with formatting
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Potential Matches", index=False)
        
        # Get the workbook and sheet
        workbook = writer.book
        worksheet = writer.sheets["Potential Matches"]
        
        # Define formats
        header_format = {'bold': True, 'text_wrap': True, 'valign': 'top', 'border': 1}
        
        # Apply conditional formatting for price differences
        last_row = len(data) + 1  # +1 for header
        
        # Create proper ColorScaleRule objects
        # Color scale for price difference percentage
        price_diff_rule = ColorScaleRule(
            start_type='min', start_color='00FF00',  # Green
            mid_type='percentile', mid_value=50, mid_color='FFFF00',  # Yellow
            end_type='max', end_color='FF0000'  # Red
        )
        
        # Color scale for confidence (reverse: green for high, red for low)
        confidence_rule = ColorScaleRule(
            start_type='min', start_color='FF0000',  # Red
            mid_type='percentile', mid_value=50, mid_color='FFFF00',  # Yellow
            end_type='max', end_color='00FF00'  # Green
        )
        
        # Apply the rules
        worksheet.conditional_formatting.add(f'L2:L{last_row}', price_diff_rule)
        worksheet.conditional_formatting.add(f'M2:M{last_row}', confidence_rule)
        
        # Format price columns as currency
        for col in [5, 10]:  # Adjust these column numbers based on your data structure
            for row in range(2, last_row + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.number_format = '£#,##0'
        
        # Format price difference as currency
        for row in range(2, last_row + 1):
            cell = worksheet.cell(row=row, column=11)  # Price difference column
            cell.number_format = '£#,##0'
            
            # Percentage format for diff percentage
            cell = worksheet.cell(row=row, column=12)  # Price difference percentage column
            cell.number_format = '0.0%'
        
        # Adjust column widths
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) if max_length < 50 else 50
            worksheet.column_dimensions[column].width = adjusted_width
    
    print(f"Exported {len(data)} potential matches to {filename}")
    
    # Optional: Sort and group matches for analysis
    # Create a second tab with price difference analysis
    analysis_df = df.sort_values(by=['price_diff_pct'], ascending=False)
    
    with pd.ExcelWriter(f"price_analysis_{platform1}_vs_{platform2}_{timestamp}.xlsx", engine='openpyxl') as writer:
        # Group 1: High confidence but large price differences (potential pricing issues)
        high_conf_price_diff = df[(df['confidence'] >= 90) & (df['price_diff_pct'] > 10)]
        high_conf_price_diff.to_excel(writer, sheet_name="High Conf Price Diffs", index=False)
        
        # Group 2: Perfect/near-perfect title matches
        title_matches = df[df['confidence'] == 100]
        title_matches.to_excel(writer, sheet_name="Perfect Matches", index=False)
        
        # Group 3: Price-matched items (within 5%)
        price_matched = df[df['price_diff_pct'] <= 5]
        price_matched.to_excel(writer, sheet_name="Price Matched", index=False)
        
        # Full dataset sorted by confidence
        df.sort_values(by=['confidence'], ascending=False).to_excel(writer, 
                                                                  sheet_name="All By Confidence", 
                                                                  index=False)
    
    print(f"Exported price analysis to price_analysis_{platform1}_vs_{platform2}_{timestamp}.xlsx")

async def main():
    # Create database connection
    db_url = os.environ.get('DATABASE_URL')
    engine = create_async_engine(db_url)
    async_session = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
    
    async with async_session() as session:
        matcher = ProductMatcher(session)
        
        # Load previous progress
        confirmed_matches, processed_pairs = await matcher.load_progress()
        
        while True:
            # Display main menu
            print("\n" + "="*80)
            print("PRODUCT MATCHER MENU")
            print("="*80)
            print("1. Find new matches to review")
            print("2. Show pending confirmed matches")
            print("3. Edit confirmed matches")
            print("4. Commit confirmed matches to database")
            print("5. Reset processed pairs")
            print("6. Export matches to Excel")
            print("7. Exit (without committing)")
            print("="*80)
            
            choice = input("Enter your choice (1-7): ")
            
            if choice == "1":
                # Find and review new matches
                await find_and_review_matches(matcher, confirmed_matches, processed_pairs)
            
            elif choice == "2":
                # Show matches waiting to be committed
                show_pending_matches(confirmed_matches)
            
            elif choice == "3":
                # Edit confirmed matches
                await edit_confirmed_matches(matcher, confirmed_matches)
            
            elif choice == "4":
                # Commit confirmed matches to database
                if confirmed_matches:
                    print(f"\nCommitting {len(confirmed_matches)} confirmed matches to database...")
                    merged = await matcher.merge_products(confirmed_matches)
                    print(f"Successfully merged {merged} products.")
                    # Clear confirmed matches after processing
                    confirmed_matches = []
                    await matcher.save_progress(confirmed_matches, processed_pairs)
                else:
                    print("No confirmed matches to commit.")
            elif choice == "5":
                # Reset processed pairs
                if input("Are you sure you want to reset all processed pairs? This will allow you to review previously rejected matches again. (y/n): ").lower() == 'y':
                    processed_pairs = set()
                    await matcher.save_progress(confirmed_matches, processed_pairs)
                    print("All processed pairs have been reset. You can now review all matches again.")
            elif choice == "6":
                await export_matches_to_excel(matcher)
            elif choice == "7":  # Exit (updated number)
                await matcher.save_progress(confirmed_matches, processed_pairs)
                print("Progress saved. Exiting without committing changes.")
                break
            
            else:
                print("Invalid choice. Please try again.")

if __name__ == "__main__":
    asyncio.run(main())