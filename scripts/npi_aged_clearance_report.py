#!/usr/bin/env python3
"""
NPI Aged Stock Clearance Report.

Tracks clearance progress on aged inventory (items listed >12 months before
the May 2025 baseline snapshot) against a £100,000 target.

Baseline files (fixed — do not change):
  data/reverb/reverb_listings_data-200525.csv  — 524 live listings, 20 May 2025
  data/reverb/reverb_listings_data-190725.csv  — Full Reverb dump, 19 Jul 2025
                                                  (provides created_at + Jul status)

The RIFF database is queried fresh on each run for current status.

Usage:
    source venv/bin/activate
    python scripts/npi_aged_clearance_report.py           # Full run + Excel output
    python scripts/npi_aged_clearance_report.py --dry-run # Print summary only, no file

Output:
    data/npi/npi_clearance_audit_DDMMYY.xlsx

Tabs:
    1. May Snapshot      — All 524 May listings with full journey + sale dates
    2. July Snapshot     — Same 524 items with July status + price changes
    3. Summary           — Clearance progress tables + age distribution
    4. All Cleared       — All 54 aged items cleared since May (SOLD + ENDED)
    5. Cleared May–Jul   — Cleared between 20 May and 19 Jul
    6. Cleared After Jul — Cleared after 19 Jul (per RIFF DB)
"""

import os
import sys
import ast
import argparse
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, text

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from app.core.config import Settings

# ── Constants ──────────────────────────────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

MAY_CSV  = os.path.join(BASE_DIR, 'data', 'reverb', 'reverb_listings_data-200525.csv')
JULY_CSV = os.path.join(BASE_DIR, 'data', 'reverb', 'reverb_listings_data-190725.csv')
OUTPUT_DIR = os.path.join(BASE_DIR, 'data', 'npi')

SNAPSHOT_DATE    = pd.Timestamp('2025-05-20', tz='UTC')
JULY_DATE        = pd.Timestamp('2025-07-19', tz='UTC')
AGED_CUTOFF      = SNAPSHOT_DATE - pd.DateOffset(months=12)  # before 20 May 2024
CLEARANCE_TARGET = 100_000


# ── Parsers ────────────────────────────────────────────────────────────────────

def parse_price(val):
    try:
        d = ast.literal_eval(str(val))
        return float(d.get('amount', 0)) if isinstance(d, dict) else None
    except Exception:
        return None


def parse_state(val):
    try:
        d = ast.literal_eval(str(val))
        return d.get('slug', 'unknown') if isinstance(d, dict) else str(val)
    except Exception:
        return str(val) if val and not isinstance(val, float) else 'unknown'


def parse_condition(val):
    try:
        d = ast.literal_eval(str(val))
        return d.get('display_name', '') if isinstance(d, dict) else str(val)
    except Exception:
        return ''


# ── Data loading ───────────────────────────────────────────────────────────────

def load_and_merge():
    """Load May + July CSVs, merge on reverb_listing_id."""
    print(f"Loading May snapshot: {MAY_CSV}")
    may = pd.read_csv(MAY_CSV)
    may['reverb_listing_id'] = may['reverb_listing_id'].astype(str)
    print(f"  {len(may)} listings")

    print(f"Loading July snapshot: {JULY_CSV}")
    jul = pd.read_csv(JULY_CSV)
    jul['id'] = jul['id'].astype(str)
    jul['created_dt'] = pd.to_datetime(jul['created_at'], utc=True, errors='coerce')
    jul['jul_state'] = jul['state'].apply(parse_state)
    jul['seller_price_gbp'] = jul['seller_price'].apply(parse_price)
    jul['condition_display'] = jul['condition'].apply(parse_condition)
    print(f"  {len(jul)} listings (created_at range: {jul['created_dt'].min().date()} → {jul['created_dt'].max().date()})")

    merged = may.merge(
        jul[['id', 'title', 'make', 'model', 'created_dt',
             'jul_state', 'seller_price_gbp', 'condition_display', 'published_at']],
        left_on='reverb_listing_id', right_on='id', how='left'
    )
    merged['age_months_at_may'] = (
        (SNAPSHOT_DATE - merged['created_dt']) / pd.Timedelta(days=30.44)
    ).round(1)
    merged['aged_flag'] = merged['created_dt'] < AGED_CUTOFF

    matched = merged['created_dt'].notna().sum()
    print(f"  Matched created_at: {matched} of {len(merged)}")
    print(f"  Aged items (>12m on 20 May 2025): {merged['aged_flag'].sum()}")
    return merged


def load_riff_data(listing_ids):
    """Query RIFF DB for current reverb_state and order/sale dates."""
    settings = Settings()
    db_url = (str(settings.DATABASE_URL)
              .replace('postgresql+asyncpg', 'postgresql+psycopg2')
              .replace('asyncpg', 'psycopg2'))
    engine = create_engine(db_url)

    with engine.connect() as conn:
        riff_rows = conn.execute(text('''
            SELECT rl.reverb_listing_id, rl.reverb_state, p.is_sold, p.status
            FROM reverb_listings rl
            JOIN platform_common pc ON pc.id = rl.platform_id
            JOIN products p ON p.id = pc.product_id
            WHERE rl.reverb_listing_id = ANY(:ids)
        '''), {'ids': listing_ids}).fetchall()

        order_rows = conn.execute(text('''
            SELECT reverb_listing_id, paid_at, amount_product, total_amount, order_number
            FROM reverb_orders
            WHERE reverb_listing_id = ANY(:ids) AND paid_at IS NOT NULL
            ORDER BY paid_at DESC
        '''), {'ids': listing_ids}).fetchall()

    riff_status = {}
    for row in riff_rows:
        lid, rstate, is_sold = str(row[0]), row[1] or 'unknown', row[2]
        if is_sold or rstate == 'sold':
            riff_status[lid] = 'sold'
        elif rstate == 'ended':
            riff_status[lid] = 'ended'
        elif rstate == 'live':
            riff_status[lid] = 'live'
        else:
            riff_status[lid] = rstate

    order_data = {}
    for row in order_rows:
        lid = str(row[0])
        if lid not in order_data:
            order_data[lid] = {
                'paid_at':       row[1],
                'sale_price_gbp': float(row[2]) if row[2] else None,
                'total_amount':  float(row[3]) if row[3] else None,
                'order_number':  str(row[4]),
            }

    print(f"  RIFF DB: {len(riff_status)} listings matched, {len(order_data)} orders with paid_at")
    return riff_status, order_data


def classify_journey(row):
    js = row['jul_state']
    rs = row['riff_state']
    if js == 'sold':
        return 'SOLD by July'
    elif js == 'ended':
        return 'ENDED by July'
    elif js == 'live':
        if rs == 'sold':   return 'SOLD after July'
        elif rs == 'ended': return 'ENDED after July'
        elif rs == 'live':  return 'STILL LIVE'
        else:               return 'LIVE Jul / not in DB'
    else:
        if rs == 'sold':   return 'SOLD (not in Jul)'
        elif rs == 'ended': return 'ENDED (not in Jul)'
        elif rs == 'live':  return 'STILL LIVE'
        else:               return 'NOT TRACKED'


# ── Console report ─────────────────────────────────────────────────────────────

def print_report(aged):
    cleared_sold  = aged[aged['journey'].str.startswith('SOLD')]['list_price'].sum()
    cleared_ended = aged[aged['journey'].str.startswith('ENDED')]['list_price'].sum()
    cleared_total = cleared_sold + cleared_ended
    pct = (cleared_total / CLEARANCE_TARGET) * 100

    journey_order = [
        'SOLD by July', 'SOLD after July',
        'ENDED by July', 'ENDED after July',
        'STILL LIVE', 'LIVE Jul / not in DB', 'NOT TRACKED',
    ]

    print()
    print('=' * 64)
    print('  NPI AGED STOCK CLEARANCE — FULL JOURNEY')
    print(f'  171 items aged >12m on 20 May 2025  |  Pool: £499,172')
    print(f'  Report: {datetime.now().strftime("%d %B %Y")}')
    print('=' * 64)
    print(f'  {"Journey":<35} {"Count":>5} {"Value (£)":>12}')
    print(f'  {"-"*35} {"-"*5} {"-"*12}')
    for j in journey_order:
        grp = aged[aged['journey'] == j]
        if grp.empty:
            continue
        print(f'  {j:<35} {len(grp):>5} {grp["list_price"].sum():>12,.0f}')
    print(f'  {"-"*35} {"-"*5} {"-"*12}')
    print(f'  {"TOTAL SOLD":<35} {aged[aged["journey"].str.contains("SOLD")].shape[0]:>5} {cleared_sold:>12,.0f}')
    print(f'  {"TOTAL ENDED":<35} {aged[aged["journey"].str.contains("ENDED")].shape[0]:>5} {cleared_ended:>12,.0f}')
    print(f'  {"CLEARED (sold+ended)":<35} {"":>5} {cleared_total:>12,.0f}')
    print()
    print(f'  TARGET:    £{CLEARANCE_TARGET:>10,}')
    print(f'  CLEARED:   £{cleared_total:>10,.0f}  ({pct:.1f}%)')
    print(f'  REMAINING: £{max(0, CLEARANCE_TARGET - cleared_total):>10,.0f}')
    bar_len = 40
    filled  = int(bar_len * min(pct, 100) / 100)
    bar     = '█' * filled + '░' * (bar_len - filled)
    print(f'  [{bar}] {pct:.1f}%')
    print('=' * 64)


# ── Excel builder ──────────────────────────────────────────────────────────────

# Shared styles
_HDR_DARK  = PatternFill('solid', fgColor='1F4E79')
_HDR_MID   = PatternFill('solid', fgColor='2E75B6')
_SUBHDR    = PatternFill('solid', fgColor='D6E4F0')
_SOLD_FILL = PatternFill('solid', fgColor='E2EFDA')
_ENDD_FILL = PatternFill('solid', fgColor='FFF2CC')
_LIVE_FILL = PatternFill('solid', fgColor='FFFFFF')
_NOTD_FILL = PatternFill('solid', fgColor='FCE4D6')
_THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin'),
)
_CENTER = Alignment(horizontal='center', vertical='center')
_TOP    = Alignment(vertical='top')

JOURNEY_FILL_MAP = {
    'SOLD by July':     _SOLD_FILL,
    'SOLD after July':  _SOLD_FILL,
    'SOLD (not in Jul)': _SOLD_FILL,
    'ENDED by July':    _ENDD_FILL,
    'ENDED after July': _ENDD_FILL,
    'ENDED (not in Jul)': _ENDD_FILL,
    'STILL LIVE':       _LIVE_FILL,
    'LIVE Jul / not in DB': _NOTD_FILL,
    'NOT TRACKED':      _NOTD_FILL,
}

STATE_FILL_MAP = {
    'sold': _SOLD_FILL,
    'ended': _ENDD_FILL,
    'live': _LIVE_FILL,
}


def _autofit(ws, min_w=8, max_w=55):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[col_letter].width = min(max_w, max(min_w, max_len + 2))


def _write_df(ws, df, start_row, fill_col=None, fill_map=None):
    """Write a DataFrame to a worksheet starting at start_row."""
    # Header
    for c_idx, col in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=c_idx, value=col)
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = _HDR_MID
        cell.alignment = _CENTER
        cell.border    = _THIN

    # Data rows
    for r_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        journey_val = None
        for c_idx, val in enumerate(row, 1):
            if isinstance(val, pd.Timestamp):
                val = val.strftime('%Y-%m-%d')
            elif isinstance(val, float) and pd.isna(val):
                val = ''
            cell           = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border    = _THIN
            cell.alignment = _TOP
            if fill_col and df.columns[c_idx - 1] == fill_col:
                journey_val = val

        # Row highlight
        if fill_map and journey_val is not None:
            row_fill = fill_map.get(journey_val)
            if row_fill:
                for c_idx in range(1, len(df.columns) + 1):
                    ws.cell(row=r_idx, column=c_idx).fill = row_fill


def _totals_row(ws, row_num, n_cols, label_col=1, val_col=None, value=None, val_col2=None, value2=None):
    for c in range(1, n_cols + 1):
        cell       = ws.cell(row=row_num, column=c)
        cell.fill  = _SUBHDR
        cell.border = _THIN
    ws.cell(row=row_num, column=label_col).value = 'TOTAL'
    ws.cell(row=row_num, column=label_col).font  = Font(bold=True)
    if val_col and value is not None:
        ws.cell(row=row_num, column=val_col).value = value
        ws.cell(row=row_num, column=val_col).font  = Font(bold=True)
    if val_col2 and value2:
        ws.cell(row=row_num, column=val_col2).value = value2
        ws.cell(row=row_num, column=val_col2).font  = Font(bold=True)


def _banner(ws, row, text, color='1F4E79', n_cols=14):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font  = Font(bold=True, size=12, color='FFFFFF')
    cell.fill  = PatternFill('solid', fgColor=color)
    ws.row_dimensions[row].height = 22
    if n_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)


def make_cleared_df(subset):
    return pd.DataFrame({
        'Reverb Listing ID':   subset['reverb_listing_id'],
        'Title':               subset['title'],
        'Brand':               subset['make'],
        'Model':               subset['model'],
        'Reverb Created At':   subset['created_dt'].dt.strftime('%Y-%m-%d').where(subset['created_dt'].notna(), ''),
        'Age at May (months)': subset['age_months_at_may'],
        'List Price GBP (May)': subset['list_price'],
        'Status May':          'live',
        'Status Jul':          subset['jul_state'],
        'Status Now (RIFF)':   subset['riff_state'],
        'Journey':             subset['journey'],
        'Paid At (Reverb)':    subset['paid_at'].apply(lambda x: x.strftime('%Y-%m-%d') if pd.notna(x) and x else ''),
        'Sale Price GBP':      subset['sale_price_gbp'].apply(lambda x: x if pd.notna(x) else ''),
        'Order Number':        subset['order_number'].fillna(''),
    }).sort_values('List Price GBP (May)', ascending=False).reset_index(drop=True)


def build_excel(merged, aged, out_path):
    wb = openpyxl.Workbook()

    # ── Tab 1: May Snapshot ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = '1. May Snapshot'
    n_cols = 16
    _banner(ws1, 1, f'MAY 2025 SNAPSHOT — All 524 live Reverb listings as of 20 May 2025', n_cols=n_cols)

    df1 = pd.DataFrame({
        'Reverb Listing ID':    merged['reverb_listing_id'],
        'Title':                merged['title'],
        'Brand':                merged['make'],
        'Model':                merged['model'],
        'Reverb Created At':    merged['created_dt'].dt.strftime('%Y-%m-%d').where(merged['created_dt'].notna(), ''),
        'Age at May (months)':  merged['age_months_at_may'],
        'Aged Flag (>12m)':     merged['aged_flag'].map({True: 'YES', False: ''}),
        'List Price GBP (May)': merged['list_price'],
        'Condition':            merged['condition_display'],
        'Status May':           'live',
        'Status Jul':           merged['jul_state'],
        'Status Now (RIFF)':    merged['riff_state'],
        'Journey':              merged['journey'],
        'Paid At (Reverb)':     merged['paid_at'].apply(lambda x: x.strftime('%Y-%m-%d') if pd.notna(x) and x else ''),
        'Sale Price GBP':       merged['sale_price_gbp'].apply(lambda x: x if pd.notna(x) else ''),
        'Order Number':         merged['order_number'].fillna(''),
    }).sort_values('List Price GBP (May)', ascending=False).reset_index(drop=True)

    _write_df(ws1, df1, start_row=2, fill_col='Journey', fill_map=JOURNEY_FILL_MAP)
    _autofit(ws1)
    ws1.freeze_panes = 'A3'

    # ── Tab 2: July Snapshot ─────────────────────────────────────────────────
    ws2 = wb.create_sheet('2. July Snapshot')
    _banner(ws2, 1, '524 May listings cross-referenced with Reverb API export of 19 Jul 2025', n_cols=13)

    df2 = pd.DataFrame({
        'Reverb Listing ID':     merged['reverb_listing_id'],
        'Title':                 merged['title'],
        'Brand':                 merged['make'],
        'Model':                 merged['model'],
        'Reverb Created At':     merged['created_dt'].dt.strftime('%Y-%m-%d').where(merged['created_dt'].notna(), ''),
        'Age at May (months)':   merged['age_months_at_may'],
        'Aged Flag (>12m)':      merged['aged_flag'].map({True: 'YES', False: ''}),
        'List Price GBP (May)':  merged['list_price'],
        'Seller Price GBP (Jul)': merged['seller_price_gbp'].apply(lambda x: x if pd.notna(x) else ''),
        'Price Change (£)':       (merged['seller_price_gbp'] - merged['list_price']).apply(
                                    lambda x: round(x, 0) if pd.notna(x) else ''),
        'Status May':            'live',
        'Status Jul':            merged['jul_state'],
        'Published At (Jul)':    merged['published_at'].fillna(''),
    }).sort_values('List Price GBP (May)', ascending=False).reset_index(drop=True)

    _write_df(ws2, df2, start_row=2, fill_col='Status Jul', fill_map=STATE_FILL_MAP)
    _autofit(ws2)
    ws2.freeze_panes = 'A3'

    # ── Tab 3: Summary ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet('3. Summary')
    pool_val      = aged['list_price'].sum()
    cleared_sold  = aged[aged['journey'].str.startswith('SOLD')]['list_price'].sum()
    cleared_ended = aged[aged['journey'].str.startswith('ENDED')]['list_price'].sum()
    cleared_total = cleared_sold + cleared_ended
    pct           = (cleared_total / CLEARANCE_TARGET) * 100

    ws3['A1'] = f'NPI Aged Stock Clearance Report — {datetime.now().strftime("%d %B %Y")}'
    ws3['A1'].font = Font(bold=True, size=14)
    ws3.merge_cells('A1:F1')
    ws3.row_dimensions[1].height = 24

    def _summary_block(ws, start_row, title, headers, data_rows, totals_row=None, color='1F4E79'):
        # Title
        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=start_row, end_column=len(headers))
        cell = ws.cell(row=start_row, column=1, value=title)
        cell.font  = Font(bold=True, size=11, color='FFFFFF')
        cell.fill  = PatternFill('solid', fgColor=color)
        ws.row_dimensions[start_row].height = 18
        r = start_row + 1
        # Headers
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = _HDR_MID; cell.alignment = _CENTER; cell.border = _THIN
        r += 1
        # Data
        for vals in data_rows:
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border    = _THIN
                cell.alignment = _CENTER
            r += 1
        # Totals
        if totals_row:
            for c, v in enumerate(totals_row, 1):
                cell       = ws.cell(row=r, column=c, value=v)
                cell.font  = Font(bold=True)
                cell.fill  = _SUBHDR
                cell.border = _THIN
                cell.alignment = _CENTER
            r += 1
        return r + 1

    journey_order = ['SOLD by July','SOLD after July','ENDED by July','ENDED after July',
                     'STILL LIVE','LIVE Jul / not in DB','NOT TRACKED']
    journey_rows  = [(j, len(aged[aged['journey']==j]),
                      f'£{aged[aged["journey"]==j]["list_price"].sum():,.0f}')
                     for j in journey_order if not aged[aged['journey']==j].empty]

    r = 3
    r = _summary_block(ws3, r,
        f'AGED STOCK JOURNEY  |  171 items aged >12 months on 20 May 2025  |  Pool: £{pool_val:,.0f}',
        ['Journey', 'Count', 'Value (£)'],
        journey_rows,
        ['TOTAL', 171, f'£{pool_val:,.0f}'],
    )
    r = _summary_block(ws3, r,
        f'CLEARANCE PROGRESS vs £{CLEARANCE_TARGET:,} TARGET',
        ['Type', 'Count', 'Value (£)'],
        [
            ['SOLD (via Reverb)',      aged[aged['journey'].str.startswith('SOLD')].shape[0],  f'£{cleared_sold:,.0f}'],
            ['ENDED (sold elsewhere)', aged[aged['journey'].str.startswith('ENDED')].shape[0], f'£{cleared_ended:,.0f}'],
        ],
        ['TOTAL CLEARED',
         aged[aged['journey'].str.contains('SOLD|ENDED')].shape[0],
         f'£{cleared_total:,.0f}  ({pct:.1f}%)'],
        color='375623',
    )

    age_buckets = [(0,3,'0–3 months'),(3,6,'3–6 months'),(6,12,'6–12 months'),
                   (12,24,'1–2 years'),(24,36,'2–3 years'),(36,999,'3+ years')]
    age_rows = [(label, len(merged[(merged['age_months_at_may']>=lo) & (merged['age_months_at_may']<hi)]),
                 f'£{merged[(merged["age_months_at_may"]>=lo) & (merged["age_months_at_may"]<hi)]["list_price"].sum():,.0f}')
                for lo, hi, label in age_buckets]
    _summary_block(ws3, r,
        'AGE DISTRIBUTION — All 524 May listings',
        ['Age Bucket', 'Items', 'Value (£)'],
        age_rows,
        ['TOTAL', 524, f'£{merged["list_price"].sum():,.0f}'],
        color='7B3F00',
    )
    ws3.column_dimensions['A'].width = 52
    ws3.column_dimensions['B'].width = 10
    ws3.column_dimensions['C'].width = 22

    # ── Cleared item tabs helper ──────────────────────────────────────────────
    def cleared_tab(ws, subtitle, color, subset):
        n = len(make_cleared_df(subset).columns)
        _banner(ws, 1, subtitle, color=color, n_cols=n)
        df = make_cleared_df(subset)
        _write_df(ws, df, start_row=2, fill_col='Journey', fill_map=JOURNEY_FILL_MAP)
        _autofit(ws)
        ws.freeze_panes = 'A3'
        tr = len(df) + 3
        _totals_row(ws, tr, n_cols=n, label_col=1,
                    val_col=7,  value=df['List Price GBP (May)'].sum(),
                    val_col2=13, value2=df['Sale Price GBP'].apply(
                        lambda x: x if isinstance(x, (int, float)) else 0).sum() or None)

    # ── Tab 4: All Cleared Since May ─────────────────────────────────────────
    ws4 = wb.create_sheet('4. All Cleared Since May')
    cleared_aged = aged[aged['journey'].str.contains('SOLD|ENDED')]
    cleared_tab(ws4,
        f'ALL CLEARED AGED ITEMS since 20 May 2025 — {len(cleared_aged)} items  |  Report: {datetime.now().strftime("%d %b %Y")}',
        '1F4E79', cleared_aged)

    # ── Tab 5: Cleared May–July ───────────────────────────────────────────────
    ws5 = wb.create_sheet('5. Cleared May–July')
    mayjul = aged[aged['journey'].isin(['SOLD by July', 'ENDED by July'])]
    cleared_tab(ws5,
        f'CLEARED between 20 May 2025 and 19 Jul 2025 — {len(mayjul)} items',
        '375623', mayjul)

    # ── Tab 6: Cleared After July ─────────────────────────────────────────────
    ws6 = wb.create_sheet('6. Cleared After July')
    aftrjul = aged[aged['journey'].isin(['SOLD after July', 'ENDED after July'])]
    cleared_tab(ws6,
        f'CLEARED after 19 Jul 2025 (live in July, now SOLD/ENDED in RIFF) — {len(aftrjul)} items',
        '7B3F00', aftrjul)

    wb.save(out_path)


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='NPI aged stock clearance report')
    parser.add_argument('--dry-run', action='store_true',
                        help='Print summary to console only, no Excel written')
    args = parser.parse_args()

    print('\n── NPI Aged Stock Clearance Report ─────────────────────────')
    merged = load_and_merge()

    print('\nQuerying RIFF database...')
    all_ids = merged['reverb_listing_id'].tolist()
    riff_status, order_data = load_riff_data(all_ids)

    merged['riff_state']    = merged['reverb_listing_id'].map(riff_status).fillna('not_in_db')
    merged['paid_at']       = merged['reverb_listing_id'].map(lambda x: order_data.get(x, {}).get('paid_at'))
    merged['sale_price_gbp'] = merged['reverb_listing_id'].map(lambda x: order_data.get(x, {}).get('sale_price_gbp'))
    merged['order_number']  = merged['reverb_listing_id'].map(lambda x: order_data.get(x, {}).get('order_number'))
    merged['journey']       = merged.apply(classify_journey, axis=1)

    aged = merged[merged['aged_flag']].copy()
    print_report(aged)

    if not args.dry_run:
        date_str = datetime.now().strftime('%d%m%y')
        out_path = os.path.join(OUTPUT_DIR, f'npi_clearance_audit_{date_str}.xlsx')
        print(f'\nBuilding Excel → {out_path}')
        build_excel(merged, aged, out_path)
        print(f'Done. {len(aged[aged["journey"].str.contains("SOLD|ENDED")])} cleared items across 6 tabs.')
    else:
        print('\n[DRY RUN] No Excel written.')


if __name__ == '__main__':
    main()
